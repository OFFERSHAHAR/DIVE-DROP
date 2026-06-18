#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
מערכת ניתוח דפי חשבון בנק אוטומטית
Bank Statement Analyzer - Hebrew Version
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json
import sys

class BankStatementAnalyzer:
    """מנתח דפי חשבון בנק"""
    
    def __init__(self, excel_file):
        """אתחול המנתח"""
        self.excel_file = excel_file
        self.df = None
        self.categories_stats = {}
        
    def load_data(self, sheet_name='גיליון1', header_row=4):
        """טעינת נתונים מ-Excel"""
        df = pd.read_excel(self.excel_file, sheet_name=sheet_name, header=header_row)
        
        # ניקוי שמות עמודות
        df.columns = ['תאריך', 'הפעולה', 'פרטים', 'אסמכתא', 'חובה', 'זכות', 
                      'יתרה', 'תאריך_ערך', 'לטובת', 'עבור']
        
        # ניקוי נתונים
        df = df.dropna(subset=['תאריך'])
        df['תאריך'] = pd.to_datetime(df['תאריך'])
        df['חובה'] = pd.to_numeric(df['חובה'], errors='coerce').fillna(0)
        df['זכות'] = pd.to_numeric(df['זכות'], errors='coerce').fillna(0)
        df['יתרה'] = pd.to_numeric(df['יתרה'], errors='coerce').fillna(0)
        
        self.df = df
        return self
    
    def categorize(self):
        """סיווג תנועות לקטגוריות"""
        def categorize_transaction(row):
            operation = str(row['הפעולה']).lower()
            details = str(row['פרטים']).lower() if pd.notna(row['פרטים']) else ''
            
            if 'עמלה' in operation or 'עמלה' in details:
                return 'עמלות בנק'
            if 'דירקט' in operation:
                return 'הוראות קבע'
            if 'שיק' in operation.lower():
                return 'צ\'קים'
            if 'העברה' in operation or 'bit' in operation:
                return 'העברות כסף'
            if 'משיכה' in operation or 'מזומן' in operation or 'מזומן' in details:
                return 'משיכת מזומן'
            if any(word in operation or word in details for word in ['הוט', 'רשום', 'צרוך']):
                return 'הוצאות קבועות'
            if row['זכות'] > 0 and row['חובה'] == 0:
                return 'הכנסות'
            return 'הוצאות משתנות'
        
        self.df['קטגוריה'] = self.df.apply(categorize_transaction, axis=1)
        return self
    
    def calculate_stats(self):
        """חישוב סטטיסטיקות"""
        for category in self.df['קטגוריה'].unique():
            cat_data = self.df[self.df['קטגוריה'] == category]
            self.categories_stats[category] = {
                'count': len(cat_data),
                'debit': cat_data['חובה'].sum(),
                'credit': cat_data['זכות'].sum(),
                'net': cat_data['זכות'].sum() - cat_data['חובה'].sum()
            }
        return self
    
    def export_excel(self, output_file):
        """ייצוא לקובץ Excel"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # הגדרות עיצוב
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # גיליון סיכום
        ws = wb.create_sheet('סיכום כללי')
        ws['A1'] = 'סיכום דפי חשבון בנק'
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws['A3'] = 'קטגוריה'
        ws['B3'] = 'מספר תנועות'
        ws['C3'] = 'סה"כ חובה'
        ws['D3'] = 'סה"כ זכות'
        
        for col in range(1, 5):
            ws.cell(row=3, column=col).fill = header_fill
            ws.cell(row=3, column=col).font = header_font
        
        row = 4
        for cat, stats in self.categories_stats.items():
            ws[f'A{row}'] = cat
            ws[f'B{row}'] = stats['count']
            ws[f'C{row}'] = stats['debit']
            ws[f'D{row}'] = stats['credit']
            row += 1
        
        wb.save(output_file)
        print(f"✓ Excel exported: {output_file}")
        return self
    
    def generate_report(self):
        """יצירת דוח טקסט"""
        print(f"\n{'='*60}")
        print(f"דוח ניתוח דפי חשבון בנק")
        print(f"{'='*60}\n")
        
        total_debit = self.df['חובה'].sum()
        total_credit = self.df['זכות'].sum()
        
        print(f"סה\"כ תנועות: {len(self.df)}")
        print(f"סה\"כ הוצאות: ₪ {total_debit:,.2f}")
        print(f"סה\"כ הכנסות: ₪ {total_credit:,.2f}")
        print(f"סל\"ד: ₪ {total_credit - total_debit:,.2f}\n")
        
        print("סיכום לפי קטגוריה:")
        print(f"{'קטגוריה':<20} {'תנועות':<10} {'חובה':<15} {'זכות':<15}")
        print("-" * 60)
        
        for cat, stats in sorted(self.categories_stats.items()):
            print(f"{cat:<20} {stats['count']:<10} ₪{stats['debit']:<14,.2f} ₪{stats['credit']:<14,.2f}")
        
        return self

def main():
    """פונקציה ראשית"""
    if len(sys.argv) < 2:
        print("שימוש: python bank_statement_analyzer.py <excel_file> [output_dir]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else '.'
    
    try:
        analyzer = BankStatementAnalyzer(excel_file)
        analyzer.load_data().categorize().calculate_stats()
        
        analyzer.generate_report()
        
        output_excel = f"{output_dir}/ניתוח_דפי_חשבון.xlsx"
        analyzer.export_excel(output_excel)
        
    except Exception as e:
        print(f"שגיאה: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
